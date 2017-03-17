import requests 
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook

def initList ():
	#initialitze table with values 
	ws.cell(row=1, column = 1).value = "Event Name"
	ws.cell(row=1, column = 2).value = "Date"
	ws.cell(row=1, column = 3).value = "Price"
	ws.cell(row=1, column = 4).value = "Line Up"
	ws.cell(row=1, column = 5).value = "Bio"
	ws.cell(row=1, column = 6).value = "Promoter"
	ws.cell(row=1, column = 7).value = "Club"
	ws.cell(row=1, column = 8).value = "Picture"
	ws.cell(row=1, column = 9).value = "Event Admin"
	ws.cell(row=1, column = 10).value = "Promo Link 1"
	ws.cell(row=1, column = 11).value = "Promo Link 2"
	ws.cell(row=1, column = 12).value = "Promo Link 3"
	ws.cell(row=1, column = 13).value = "Event Link"
	ws.cell(row=1, column = 14).value = " "

def MastherMethod (eventlinks):
	y=1 #row counter
	for event in eventlinks:
		try: 
			y+=1 
			r = requests.get(event)
			soup = BeautifulSoup(r.content, 'lxml')

			ws.cell(row=y, column = 13).value = event

			eventname = soup.find('h1').string 
			ws.cell(row=y, column=1).value = eventname

			new_soup = soup.find('div', {'id': 'event-item'})

			#get line up and bio
			helper = new_soup.find_all('p')
			try: 
				line_up1 = helper[0].get_text()
				line_up2 = line_up1.encode('utf8')
				ws.cell(row=y, column=4).value = line_up2
				bio1 = helper[1].get_text()
				bio2 = bio1.encode('utf8')
				ws.cell(row=y, column=5).value = bio2
			except: 
				pass

			#get flyer picture
			try: 
				helper2 = new_soup.find('div', {'class': 'flyer'})
				flyer_dirty= helper2.find('a', href=True)
				flyer = "https://www.residentadvisor.net" + flyer_dirty['href']
				ws.cell(row=y, column=8).value = flyer
			except: 
				pass

			#get all the links, first is event admin, then update event, then promo links!
			helper3 = new_soup.find('div', {'class': 'clearfix right'})

			links = helper3.find('div', {'class': 'links'})
			links2 = links.find_all('a', href=True)
			try: 
				admin = links2[0].get_text()
				ws.cell(row=y, column = 9).value = admin
			except:
				ws.cell(row=y, column = 9).value = " "

			try: 
				promolink1 = links2[2]['href']
				# promolink2 = promolink1['href']
				ws.cell(row=y, column = 10).value = promolink1
			except: 
				ws.cell(row=y, column = 10).value = " "


			try: 
				promolink2 = links2[3]['href']
				ws.cell(row=y, column = 11).value = promolink2
			except: 
				ws.cell(row=y, column = 11).value = " " 
			try: 
				promolink3 = links2[4]['href']
				ws.cell(row=y, column = 12).value = promolink3
			except: 
				ws.cell(row=y, column = 12).value = " "

			test = soup.find('ul', {'class': 'clearfix'})
			test2 = test.find_all('a', href=True)

			for item in test2:
				try:
					if "events.aspx" in item['href']:
						#date 
						ws.cell(row=y, column=2).value = item.get_text()
				except: 
					ws.cell(row=y, column=2).value = " "
				try:
					if "club.aspx" in item['href']:
						ws.cell(row=y, column=7).value = item.get_text()
				except:
					ws.cell(row=y, column=7).value = " "

			#get costs for event
			li = test.find_all('li')
			for item in li:
				try: 
					if "Cost" in item.get_text():
						cost_dirty = item.get_text()
						cost_clean = cost_dirty.replace('Cost /', '')
						ws.cell(row=y, column = 3).value = cost_clean
				except: 
					ws.cell(row=y, column = 3).value = " "
				
				try: 
					if "Promoters /" in item.get_text():
						prom_dirty = item.get_text()
						prom_clean1 = prom_dirty.replace('Promoters /', '')
						prom_clean2 = prom_clean1.encode('utf8')
						ws.cell(row=y, column = 6).value = prom_clean2
			 	except:
					ws.cell(row=y, column = 6).value = " "

			ws.cell(row=y, column = 14).value = " "

		except: 
			pass

##################	Main Method ##################
wb = Workbook()
filename1 = 'germany_eventlist.xlsx'
ws = wb.active
ws.title = 'germany_eventlist'

file_path = '/Users/nequalstim/Desktop/bookya/germany'

initList()

eventlist = [line.rstrip('\n') for line in open(os.path.join(file_path, 'germany_event.txt'), 'r')]

MastherMethod(eventlist)

wb.save(filename = filename1)
