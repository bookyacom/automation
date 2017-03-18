import requests 
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook

def initList ():
	#initialitze table with values 
	ws.cell(row=1, column = 1).value = "Name Promoter"
	ws.cell(row=1, column = 2).value = "Location"
	ws.cell(row=1, column = 3).value = "RA_link"
	ws.cell(row=1, column = 4).value = "Genre"
	ws.cell(row=1, column = 5).value = "Website"
	ws.cell(row=1, column = 6).value = "Picture"
	ws.cell(row=1, column = 7).value = "Bio"
	ws.cell(row=1, column = 8).value = "Contact Person"
	ws.cell(row=1, column = 9).value = "Email"

#get all the venue links and save them
def getAllLinks (RA_id, list_end):

	url = "https://www.residentadvisor.net/promoters.aspx?ai=" + str(RA_id)
	r = requests.get(url)
	soup = BeautifulSoup(r.content, "lxml")

	for a in soup.find_all('a', href=True):
		if promoter in a['href']:
	   		prom_links.append("https://www.residentadvisor.net" + a['href'])
	   	#marking the end of the list
	   	if str(list_end) in a['href']:
	   		break

#iterate over venue links and clean out the ones which didn't have a listing in 2016
def filterProms():
	cut_off = '2016'
	cut_off2 = '2017'
	for site in prom_links:
		prom_url = site
		new_r = requests.get(prom_url)
		new_soup = BeautifulSoup(new_r.content, "lxml")

		#archive 
		archive_listing = new_soup.find('div', {'id': 'divArchiveEvents'})
		for a in archive_listing.find_all('time', datetime = True):
			if cut_off in a['datetime']:
				prom_links_dirty.append(prom_url)
			elif cut_off2 in a['datetime']:
				prom_links_dirty.append(prom_url)
	
	prom_links_filtered = sorted(list(set(prom_links_dirty)))

	promfile = open(os.path.join(file_path, 'COUNTRY_promlinks.txt'), 'w')

	for promlink in prom_links_filtered:
		promfile.write(promlink + "\n")

	promfile.close()

	return prom_links_filtered

def masterMethod(links):

	y = 1
	for item in links:
		y += 1
		ven_url = item 
		the_r = requests.get(ven_url)
		the_soup = BeautifulSoup(the_r.content, "lxml")

		ws.cell(row = y, column = 3).value = ven_url

		# print out sorted (nach residentadvisorlink) list of promoters
		promotername = the_soup.find('h1').string
		promotername = promotername.encode('utf8')
		ws.cell(row = y, column = 1).value = promotername

		# print bio to file
		bio_dirty = the_soup.find('div', {'style': 'padding:16px 32px 32px 0;'})
		bio = bio_dirty.get_text()
		bio2 = bio.encode('utf8')
		ws.cell(row = y, column = 7).value = bio2

		#get picture links
		pic = the_soup.find_all('img', src = True)
		promoter = pic[1]
		if "promoter" in promoter['src']:
			picture = "https://www.residentadvisor.net" + promoter['src']
			ws.cell(row = y, column = 6).value = picture 
		else: 
			ws.cell(row = y, column = 6).value = "no picture"


		#Get Website and mail address

		newsoup = the_soup.find('ul', {'class': 'clearfix'})

		sites = newsoup.find_all('a', href= True)
		homepage = sites[1]
		if "Website" in homepage: 
			site = homepage['href']
			ws.cell(row = y, column = 5).value = site
		else:
			ws.cell(row = y, column = 5).value = "no website"

		try:
			mailpage = sites[2]
			if "Email" in mailpage:
				email_dirt = mailpage['href']
				email_clean = email_dirt.replace("mailto:", "")
				ws.cell(row = y, column = 9).value = email_clean
			else: 
				ws.cell(row = y, column = 9).value = "no email"
		except: 
			ws.cell(row = y, column = 9).value = "no email"
		
		#print location to file
		loc_dirty = sites[0]
		loc_clean = loc_dirty['title']
		ws.cell(row = y, column = 2).value = loc_clean

		#extract all the event links from venues
		event = "event.aspx?"
		for link in the_soup.find_all('a', href=True):
			if event in link['href']:
				e_link = "https://www.residentadvisor.net" + link['href']
				prom_event_links.write(e_link + "\n")

		# put a space in empty cells for nicer formatting in excel 
		placeholders = [4, 8]
		for column in placeholders:
			ws.cell(row = y, column = column).value = " "

###############MAIN ROUTINE################

# set up workbook
wb = Workbook()
filename1 = 'COUNTRY_prom.xlsx'
ws = wb.active
ws.title = 'COUNTRY'

file_path = '/Users/nequalstim/Desktop/bookya/COUNTRY'
prom_event_links = open(os.path.join(file_path, 'COUNTRY_promeventlinks.txt'), 'w')

prom_links = []
promoter = "promoter.aspx?"
prom_links_dirty = []

RA_ids = []

list_end = []

initList()
# get all the promoter links from all of COUNTRY and save them to COUNTRY_promlinks
for x, y in zip(RA_ids, list_end):
	try:	
		getAllLinks(x, y)
	except:
		pass

#filter out all the venues that didn't have an event in 2016,2017
prom_filter = filterProms()
# promfilter2 = [line.rstrip('\n') for line in open(os.path.join(file_path, 'COUNTRY_promlinks.txt'), 'r')]


#extract all the data from the venue pages and write it to excel file 
#additionally write the event links of each club to seperate file 
masterMethod(prom_filter)

wb.save(filename = filename1)
prom_event_links.close()