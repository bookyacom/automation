#TODO only want those aritsts and venues that are already in the DB
import json
import requests 
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook
import re

def initList ():
	#initialitze table with values 
	ws.cell(row=1, column = 1).value = "display_name"
	ws.cell(row=1, column = 2).value = "profile_photo"
	ws.cell(row=1, column = 3).value = "email"
	ws.cell(row=1, column = 4).value = "public_email"
	ws.cell(row=1, column = 5).value = "public_contact_number"
	ws.cell(row=1, column = 6).value = "search_artist"
	ws.cell(row=1, column = 7).value = "contact_number"
	ws.cell(row=1, column = 8).value = "bio"
	ws.cell(row=1, column = 9).value = "websites"
	ws.cell(row=1, column = 10).value = "genre_list"
	ws.cell(row=1, column = 11).value = "start_date"
	ws.cell(row=1, column = 12).value = "end_date"
	ws.cell(row=1, column = 13).value = "ticket_price"
	ws.cell(row=1, column = 14).value = "ticket_currency"
	ws.cell(row=1, column = 15).value = "contact_person"
	ws.cell(row=1, column = 16).value = "promoter_name"
	ws.cell(row=1, column = 17).value = "promoter_url"
	ws.cell(row=1, column = 18).value = "venue_name"
	ws.cell(row=1, column = 19).value = "venue_url"
	ws.cell(row=1, column = 20).value = "artist_names_comma_seperated"
	ws.cell(row=1, column = 21).value = "artist_urls_comma_seperated"
	ws.cell(row=1, column = 22).value = "facebook_page"
	ws.cell(row=1, column = 23).value = "beatport_dj"
	ws.cell(row=1, column = 24).value = "instagram"
	ws.cell(row=1, column = 25).value = "lastfm"
	ws.cell(row=1, column = 26).value = "mixcloud"
	ws.cell(row=1, column = 27).value = "partyflock"
	ws.cell(row=1, column = 28).value = "songkick"
	ws.cell(row=1, column = 29).value = "soundcloud"
	ws.cell(row=1, column = 30).value = "spotify"
	ws.cell(row=1, column = 31).value = "twitter"
	ws.cell(row=1, column = 32).value = "youtube_channel"
	ws.cell(row=1, column = 33).value = "bandsintown"
	ws.cell(row=1, column = 34).value = "Garbage lineup"
	ws.cell(row=1, column = 35).value = "link1"
	ws.cell(row=1, column = 36).value = "link2"
	ws.cell(row=1, column = 37).value = "link3"


def MastherMethod (eventlinks):
	y=1 #row counter
	api_url = "https://admin-api-test.bookya.com/admin/check?"
	for event in eventlinks:

		y+=1 
		r = requests.get(event)
		soup = BeautifulSoup(r.content, 'lxml')

		eventname = soup.find('h1').string 
		ws.cell(row=y, column=1).value = eventname

		ws.cell(row=y, column = 3).value = "noreply@bookya.com"

		new_soup = soup.find('div', {'id': 'event-item'})

		#get line up 
		helper = new_soup.find_all('p')
		try:
			helper_array_urls = [] 
			helper_array_line = [] 
			line_up1 = helper[0]
			line_up2 = line_up1.find_all('a', href=True)
			for item in line_up2:
				parameters = {"name": item, "type": "artist"}
				response = requests.get(api_url, params=parameters)
				data = response.json()
				if not data['profiles']:
					helper_array_line.append(item.get_text().encode('utf8'))
				else: 
					helper_array_urls.append(data['profiles'][0]['bookya_url'])

			ws.cell(row=y, column = 20).value = str(helper_array_line)
			ws.cell(row=y, column = 21).value = str(helper_array_urls)
		except: 
			pass

		#get bio
		try: 
			bio1 = helper[1].get_text()
			bio2 = bio1.encode('utf8')
			ws.cell(row=y, column=8).value = bio2
		except:
			pass

		#get flyer picture
		try: 
			helper2 = new_soup.find('div', {'class': 'flyer'})
			flyer_dirty= helper2.find('a', href=True)
			flyer = "https://www.residentadvisor.net" + flyer_dirty['href']
			ws.cell(row=y, column=2).value = flyer
		except: 
			ws.cell(row=y, column=2).value = " "

		#get all the links, first is event admin, then update event, then promo links!
		helper3 = new_soup.find('div', {'class': 'clearfix right'})

		links = helper3.find('div', {'class': 'links'})
		links2 = links.find_all('a', href=True)

		#not really useful, because it's the RA account name
		# try: 
		# 	promoter_name = links2[0].get_text()
		# 	print "links2[0] = " + promoter_name
		# 	ws.cell(row=y, column = 16).value = promoter_name
		# except:
		# 	ws.cell(row=y, column = 16).value = " "



		test = soup.find('ul', {'class': 'clearfix'})
		test2 = test.find_all('a', href=True)
		z=11
		for item in test2:
			try:
				if "events.aspx" in item['href']:
					#date
					ws.cell(row=y, column=z).value = item.get_text()
					z+=1
			except: 
				ws.cell(row=y, column=11).value = " "

			try:
				if "club.aspx" in item['href']:
					venue_n = item.get_text()
					parameters = {"name": venue_n, "type": "venue"}
					response = requests.get(api_url, params=parameters)
					data = response.json()
					if not data['profiles']:
						ws.cell(row=y, column=18).value = venue_n
					else:
						ws.cell(row=y, column=18).value = venue_n
						ws.cell(row=y, column=19).value = data['profiles'][0]['bookya_url']

			except:
				ws.cell(row=y, column=18).value = " "

			try:
				if "promoter.aspx" in item['href']:
					promoter = item.get_text()
					parameters = {"name": promoter, "type": "promoter"}
					response = requests.get(api_url, params=parameters)
					data = response.json()
					if not data['profiles']:
						ws.cell(row=y, column=16).value = promoter
					else: 
						ws.cell(row=y, column=16).value = promoter
						ws.cell(row=y, column=17).value = data['profiles'][0]['bookya_url']
			except:
				ws.cell(row=y, column=16).value = " "




		# here you get all the socials from the right side, not useful
		# try: 
		# 	promolink1 = links2[2]['href']
		# 	print "links2[2] = " + links2[2]['href']
		# 	if venue_n.lower() in promolink1:
		# 		ws.cell(row=y, column = 35).value = promolink1
		# 	else:
		# 		ws.cell(row=y, column = 35).value = promolink1
		# except: 
		# 	ws.cell(row=y, column = 35).value = " "


		# try: 
		# 	promolink2 = links2[3]['href']
		# 	print "links2[3] = " + links2[3]['href']
		# 	if venue_n.lower() in promolink2:
		# 		ws.cell(row=y, column = 36).value = promolink2
		# 	else:
		# 		ws.cell(row=y, column = 36).value = promolink2
		# except:
		# 	ws.cell(row=y, column = 36).value = " " 
		# try: 
		# 	promolink3 = links2[4]['href']
		# 	print "links2[4] = " + links2[4]['href']
		# 	print "---------------------------------"
		# 	if venue_n.lower() in promolink3:
		# 		ws.cell(row=y, column = 37).value = promolink3
		# 	else:
		# 		ws.cell(row=y, column = 37).value = promolink3
		# except: 
		# 	ws.cell(row=y, column = 37).value = " "		

		#get costs for event
		li = test.find_all('li')
		for item in li:
			try: 
				if "Cost" in item.get_text():
					cost_dirty = item.get_text()
					cost_clean = cost_dirty.replace('Cost /', '')
					ws.cell(row=y, column = 13).value = cost_clean
				else: 
					ws.cell(row=y, column = 13).value = " "
			except: 
				ws.cell(row=y, column = 13).value = " "
			
			try: 
				if "Promoters /" in item.get_text():
					prom_dirty = item.get_text()
					prom_clean1 = prom_dirty.replace('Promoters /', '')
					prom_clean2 = prom_clean1.encode('utf8')
					ws.cell(row=y, column = 32).value = prom_clean2
				else: 
					ws.cell(row=y, column = 32).value = " "
		 	except:
				ws.cell(row=y, column = 32).value = " "

		# put a space in empty cells for nicer formatting in excel 
		placeholders = [4, 5, 6, 7, 9, 10, 14, 15, 22, 23, 24, 25, 29, 30, 31, 33]
		for column in placeholders:
			ws.cell(row = y, column = column).value = " "

		print "-------------------------------"

##################	Main Method ##################
wb = Workbook()
filename1 = 'COUNTRY_events.xlsx'
ws = wb.active
ws.title = 'Event'

file_path = '/Users/nequalstim/Google Drive/DATABASE/EMEA/COUNTRY'

initList()

#read all the eventlinks from file outputted by promoter or venue scraper
eventlist = [line.rstrip('\n') for line in open(os.path.join(file_path, 'test_events.txt'), 'r')]

MastherMethod(eventlist)

wb.save(file_path + '/' + filename1)
