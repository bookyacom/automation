import requests 
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook

def initList ():
	#initialitze table with values 
	ws.cell(row=1, column = 1).value = "Venue Name"
	ws.cell(row=1, column = 2).value = "RA URL"
	ws.cell(row=1, column = 3).value = "Profile Pic"
	ws.cell(row=1, column = 4).value = "Location"
	ws.cell(row=1, column = 5).value = "Google Place ID"
	ws.cell(row=1, column = 6).value = "Bio"
	ws.cell(row=1, column = 7).value = "Genres"
	ws.cell(row=1, column = 8).value = "Capacity"
	ws.cell(row=1, column = 9).value = "Performance Area"
	ws.cell(row=1, column = 10).value = "Type of Venue"
	ws.cell(row=1, column = 11).value = " "
	ws.cell(row=1, column = 12).value = "Website"
	ws.cell(row=1, column = 13).value = "Facebook"
	ws.cell(row=1, column = 14).value = "Instagram"
	ws.cell(row=1, column = 15).value = "Songkick"
	ws.cell(row=1, column = 16).value = "Twitter"
	ws.cell(row=1, column = 17).value = "Youtube"
	ws.cell(row=1, column = 18).value = " "
	ws.cell(row=1, column = 19).value = "Main Contact"
	ws.cell(row=1, column = 20).value = "Email"
	ws.cell(row=1, column = 21).value = "Phone number"
	ws.cell(row=1, column = 22).value = "Address"
	ws.cell(row=1, column = 23).value = " "

#get all the venue links and save them
def getAllLinks (RA_id, list_end):

	url = "https://www.residentadvisor.net/clubs.aspx?ai=" + str(RA_id)
	r = requests.get(url)
	soup = BeautifulSoup(r.content, "lxml")

	for a in soup.find_all('a', href=True):
		if venue in a['href']:
	   		venue_links.append("https://www.residentadvisor.net" + a['href'])
	   	#marking the end of the list
	   	if str(list_end) in a['href']:
	   		break

#iterate over venue links and clean out the ones which didn't have a listing in 2016
def filterClubs():
	cut_off = '2016'
	cut_off2 = '2017'
	for site in venue_links:
		venue_url = site
		new_r = requests.get(venue_url)
		new_soup = BeautifulSoup(new_r.content, "lxml")

		#archive 
		archive_listing = new_soup.find('div', {'id': 'divArchiveEvents'})
		for a in archive_listing.find_all('time', datetime = True):
			if cut_off in a['datetime']:
				venue_links_dirty.append(venue_url)
			elif cut_off2 in a['datetime']:
				venue_links_dirty.append(venue_url)
	
	venue_links_filtered = sorted(list(set(venue_links_dirty)))

	venfile = open(os.path.join(file_path, 'INSERT_COUNTRY_ven_links.txt'), 'w')

	for venlink in venue_links_filtered:
		venfile.write(venlink + "\n")

	venfile.close()

	return venue_links_filtered


def masterMethod(vlfilter):

	y = 1
	for item in vlfilter:
		y += 1
		ven_url = item 
		the_r = requests.get(ven_url)
		the_soup = BeautifulSoup(the_r.content, "lxml")

		ws.cell(row = y, column = 2).value = ven_url

		# print out sorted (nach residentadvisorlink) list of clubs
		clubname = the_soup.find('h1').string
		clubname1 = clubname.encode('utf8')
		ws.cell(row = y, column = 1).value = clubname1

		# print bio to file
		bio_dirty = the_soup.find('div', {'style': 'padding:16px 32px 32px 0;'})
		bio = bio_dirty.get_text()
		bio2 = bio.encode('utf8')
		ws.cell(row = y, column = 6).value = bio2

		#print adresses of clubs
		try:
			address_dirty = the_soup.find('span', {'itemprop': 'street-address'})
			address = address_dirty.get_text()
			ws.cell(row = y, column = 22).value = address
		except:
			ws.cell(row = y, column = 22).value = "no address"


		#get picture links
		pic = the_soup.find_all('img', src = True)
		try:
			club = pic[1]
			if "clubs" in club['src']:
				picture = "https://www.residentadvisor.net" + club['src']
				ws.cell(row = y, column = 3).value = picture 
			else: 
				ws.cell(row = y, column = 3).value = "no picture"
		except:
			ws.cell(row = y, column = 3).value = "no picture"

		#Get Website and mail addres
		newsoup = the_soup.find('ul', {'class': 'clearfix'})

		sites = newsoup.find_all('a', href= True)
		try: 
			homepage = sites[1]
			if "Website" in homepage: 
				site = homepage['href']
				ws.cell(row = y, column = 12).value = site
			else:
				ws.cell(row = y, column = 12).value = "no website"
		except: 
			ws.cell(row = y, column = 12).value = "no website"

		try:
			mailpage = sites[3]
			if "Email" in mailpage:
				email_dirt = mailpage['href']
				email_clean = email_dirt.replace("mailto:", "")
				ws.cell(row = y, column = 20).value = email_clean
			else: 
				ws.cell(row = y, column = 20).value = "no email"
		except: 
			ws.cell(row = y, column = 20).value = "no email"
				
		#Get capacity number
		lis = newsoup.find_all('li')
		try: 
			capacity = lis[1]
			capacity_dirt = capacity.get_text()
			if "Capacity" in capacity_dirt:
				capacity_clean = capacity_dirt.replace("Capacity /", "")
				ws.cell(row = y, column = 8).value = capacity_clean
			else:
				ws.cell(row = y, column = 8).value = "no capacity"
		except:
			ws.cell(row = y, column = 8).value = "no capacity"


		#extract all the event links from venues
		event = "event.aspx?"
		for links in the_soup.find_all('a', href=True):
			if event in links['href']:
				e_link = "https://www.residentadvisor.net" + links['href']
				file_event_links.write(e_link + "\n")

		# put a space in empty cells for nicer formatting in excel 
		placeholders = [4, 5, 7, 9, 10, 11, 13, 14, 15, 16, 17, 18, 19, 21, 23]
		for column in placeholders:
			ws.cell(row = y, column = column).value = " "

###############MAIN ROUTINE################

# set up workbook
wb = Workbook()
filename1 = 'INSERT_COUNTRY.xlsx'
ws = wb.active
ws.title = 'INSERT_COUNTRY'

file_path = '/Users/nequalstim/Desktop/bookya/INSERT_COUNTRY'
file_event_links = open(os.path.join(file_path, 'INSERT_COUNTRY_event.txt'), 'w')

venue_links = []
venue = "club.aspx?"
venue_links_dirty = []

#Enter RA IDs here 
RA_ids = []

#correspondant list endings here 
list_end = []

initList()
# get all the venue links from all of Germany and save them to venue_links
for x, y in zip(RA_ids, list_end):
	getAllLinks(x, y)

#filter out all the venues that didn't have an event in 2016,2017
vlfilter = filterClubs()

#just in case, we have to run it again to extract data, we don't have to filter all the links from scratch
# vlfilter2 = [line.rstrip('\n') for line in open(os.path.join(file_path, 'INSERT_COUNTRY_ven_links.txt'), 'r')]

#extract all the data from the venue pages and write it to excel file 
#additionally write the event links of each club to seperate file 
masterMethod(vlfilter)

wb.save(filename = filename1)
file_event_links.close()