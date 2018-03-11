from openpyxl import Workbook
import requests

def event_header(ws):
    #initialitze table with values
    ws.cell(row=1, column = 1).value = "display_name"
    ws.cell(row=1, column = 2).value = "profile_photo"
    ws.cell(row=1, column = 3).value = "email"
    ws.cell(row=1, column = 4).value = "public_email"
    ws.cell(row=1, column = 5).value = "public_contact_number"
    ws.cell(row=1, column = 6).value = "search_artist"
    ws.cell(row=1, column = 7).value = "contact_number"
    ws.cell(row=1, column = 8).value = "bio"
    ws.cell(row=1, column = 9).value = "websites"
    ws.cell(row=1, column = 10).value = "genre_list"
    ws.cell(row=1, column = 11).value = "start_date"
    ws.cell(row=1, column = 12).value = "month"
    ws.cell(row=1, column = 13).value = "year"
    ws.cell(row=1, column = 14).value = "ticket_price"
    ws.cell(row=1, column = 15).value = "contact_person"
    ws.cell(row=1, column = 16).value = "promoter_name"
    ws.cell(row=1, column = 17).value = "promoter_url"
    ws.cell(row=1, column = 18).value = "CITY"
    ws.cell(row=1, column = 19).value = "COUNTRY"
    ws.cell(row=1, column = 20).value = "artist_names_comma_seperated"
    ws.cell(row=1, column = 21).value = "artist_urls_comma_seperated"
    ws.cell(row=1, column = 22).value = "facebook_page"
    ws.cell(row=1, column = 23).value = "beatport_dj"
    ws.cell(row=1, column = 24).value = "instagram"
    ws.cell(row=1, column = 25).value = "lastfm"
    ws.cell(row=1, column = 26).value = "mixcloud"
    ws.cell(row=1, column = 27).value = "partyflock"
    ws.cell(row=1, column = 28).value = "songkick"
    ws.cell(row=1, column = 29).value = "soundcloud"
    ws.cell(row=1, column = 30).value = "spotify"
    ws.cell(row=1, column = 31).value = "twitter"
    ws.cell(row=1, column = 32).value = "youtube_channel"
    ws.cell(row=1, column = 33).value = "bandsintown"
    ws.cell(row=1, column = 34).value = "garbage 1"
    ws.cell(row=1, column = 35).value = "garbage 2"

def request_db(request, type_):
    """
    Returns matchings of a request in the Bookya DB

    Arugments:
    request: name of festival, artist, event
    type_: promoter, artist, event?

    Return: 
    matchings: array of matches
    """

    api_url = "https://admin-api.bookya.com/admin/check?"

    req = request.encode('utf8')
    parameters = {"name": req, "type": type_}
    response = requests.get(api_url, params=parameters)
    data = response.json()

    return data

def write_to_file(array, fp):
    for item in array:
        fp.write(item+'\n')

    fp.close()
