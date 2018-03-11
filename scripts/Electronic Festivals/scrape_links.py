from openpyxl import Workbook
import os

from soup import get_soup
from helper import *
from site_elements import *


ef_url = 'http://www.electronic-festivals.com'

def scrape_events(links):
    """
    This method will take in links from Electronic Festivals website,
    collect all the relevant information and write them into an Excel sheet. 

    Arguments:
    links: array filled with EF links

    Return:
    None

    Side Effects:
    Create Excel File
    """

    #Init Excel
    wb = Workbook()
    filename = 'EF_events.xlsx'
    filepath = os.getcwd()
    ws = wb.active
    ws.title = 'Electronic Festivals'

    event_header(ws)

    #array that will be filled with all DJs that had multiple matches in the Bookya DB
    problem_djs = []

    for row_count, link in enumerate(links, start=2):

        site = get_soup(ef_url+link)

        middle_block = site.find('div', {'typeof': 'http://schema.org/Festival'})

        #Get Genre
        ws.cell(row=row_count, column=10).value = genre(middle_block)

        #Get start and end date
        date_festival = date(middle_block)
        ws.cell(row = row_count, column = 11).value = date_festival

        #Get name of festival
        name_ = name(middle_block)
        ws.cell(row = row_count, column = 1).value = name_

        ws.cell(row = row_count, column = 8).value = bio(middle_block)

        #needed to match promoter
        website = homepage(middle_block)
        ws.cell(row = row_count, column = 9).value = website

        #Get location of Event
        location_ = location(middle_block)
        ws.cell(row = row_count, column = 19).value = location_

        #get facebook, youtube links
        facebook, youtube = get_promo_urls(middle_block)
        ws.cell(row = row_count, column = 22).value = facebook
        ws.cell(row = row_count, column = 32).value = youtube

        #Match artists to Bookya DB, artist with multiple matches will be written to file
        artist_names, artist_urls = line_up(middle_block, problem_djs, date_festival)
        ws.cell(row = row_count, column = 20).value = ','.join(artist_names)
        ws.cell(row = row_count, column = 21).value = ','.join(artist_urls)

        #Match Event to Promoter in Bookya DB
        promoter_name, promoter_url = promoter(name_, website, location_)
        ws.cell(row = row_count, column = 16).value = promoter_name
        ws.cell(row = row_count, column = 17).value = promoter_url

        # put a space in empty cells for nicer formatting in excel
        placeholders = [2, 3, 4, 5, 6, 13, 14, 15, 18, 19, 23, 24, 25, 26, 27, 28, 29, 30, 31, 33]
        for column in placeholders:
            ws.cell(row = row_count, column = column).value = " "

    dj_fp = open(os.path.join(filepath, "DJs_multiple_matches.txt"), 'w')
    problem_djs = list(set(problem_djs))
    write_to_file(problem_djs, dj_fp)
    wb.save(filepath + '/' + filename)

