from bs4 import BeautifulSoup
from openpyxl import Workbook
from utils import *

test=True

def MastherMethod (eventlinks):
	"""
    Scrape each link's event information

	Arguments:
	eventlinks: list of links to be scraped

	Side Effects:
	problem_artists: File with DJ names that have more than 1 match in Database
    """
	y=1 #row counter

	problem_artists = []
	for event in eventlinks:

		y+=1
		r = requests.get(event)
		soup = BeautifulSoup(r.content, 'lxml')

		ws.cell(row=y, column=1).value = get_event_name(soup)

		ws.cell(row=y, column = 4).value = "noreply@bookya.com"

		new_soup = soup.find('div', {'id': 'event-item'})
		p_tags = new_soup.find_all('p')

		#get line up
		notBookya, bookya, problem_artists = line_up(p_tags, problem_artists)
		ws.cell(row=y, column = 22).value = ','.join(notBookya)
		ws.cell(row=y, column = 23).value = ','.join(bookya)

		#get bio
		bio_ = bio(p_tags)
		ws.cell(row=y, column=10).value = bio_

		#get flyer picture
		ws.cell(row=y, column=3).value = flyer_picture(new_soup)

		#dates and venue both need a_tags from top_bar
		top_bar = soup.find('ul', {'class': 'clearfix'})
		a_tags = top_bar.find_all('a', href=True)

		start_date, end_date = dates(a_tags)
		ws.cell(row=y, column=13).value = start_date
		ws.cell(row=y, column=14).value = end_date

		found_venue, venue_name = venue(a_tags)
		if (found_venue):
			ws.cell(row=y, column=21).value = venue_name
		else:
			ws.cell(row=y, column=20).value = venue_name



		#get costs for event
		li = top_bar.find_all('li')
		cost, currency = costs(li)
		ws.cell(row=y, column = 15).value = cost
		ws.cell(row=y, column = 16).value = currency

		found_prom, promoter_name = promoter(li)
		if(found_prom):
			ws.cell(row=y, column=19).value = promoter_name
		else:
			ws.cell(row=y, column=18).value = promoter_name

		# put a space in empty cells for nicer formatting in excel
		placeholders = [2, 5, 6, 7, 8, 9, 11, 12, 17, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35]
		for column in placeholders:
			ws.cell(row = y, column = column).value = " "

	#write problem_artists to file
	artistfile = open(os.path.join(file_path, 'problem_artists.txt'), 'w')
	problem_artists = list(set(problem_artists))
	for dj in problem_artists:
		artistfile.write(dj+'\n')
	artistfile.close()

##################	Main Method ##################
wb = Workbook()
ws = wb.active
ws.title = 'Event'

file_path = '/Users/nequalstim/Desktop/bookya/RA/'

set_up_event(ws)

#read all the eventlinks from file outputted by promoter or venue scraper
if(test):
	eventlist = fill_eventlist(file_path, 'test_events.txt')
	filename1 = 'Test_events.xlsx'
else:
	eventlist = fill_eventlist(file_path, 'eventlinks_master.txt')
	filename1 = 'RA_events.xlsx'


MastherMethod(eventlist)

wb.save(file_path + '/' + filename1)
