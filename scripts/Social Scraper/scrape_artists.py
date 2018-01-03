from openpyxl import Workbook
import os
from viberate import viberate
from socials import lastfm, record_labels, facebook, instagram, songkick, partyflock

def init_artist_profiles(ws):
    ws.cell(row=1, column = 1).value = 'display_name'
    ws.cell(row=1, column = 2).value = 'websites'
    ws.cell(row=1, column = 3).value = 'record_labels'
    ws.cell(row=1, column = 4).value = 'facebook_page'
    ws.cell(row=1, column = 5).value = 'beatport_dj'
    ws.cell(row=1, column = 6).value = 'instagram'
    ws.cell(row=1, column = 7).value = 'lastfm'
    ws.cell(row=1, column = 8).value = 'mixcloud'
    ws.cell(row=1, column = 9).value = 'partyflock'
    ws.cell(row=1, column = 10).value = 'songkick'
    ws.cell(row=1, column = 11).value = 'soundcloud'
    ws.cell(row=1, column = 12).value = 'spotify'
    ws.cell(row=1, column = 13).value = 'twitter'
    ws.cell(row=1, column = 14).value = 'youtube_channel'
    ws.cell(row=1, column = 15).value = 'bandsintown'


def scrape_artists(artists):
    """
    Scrape all the social information for artists

    First try on Viberate (they have good data)
    If artist not on Viberate scrape manually

    Arguments:
    artists: Array filled with artist names

    Side Effects:
    Create DJ_socials.xlsx with all info
    """
    wb = Workbook()
    filename = 'DJ_socials.xlsx'
    ws = wb.active
    ws.title = 'DJ_socials'
    file_path = '/Users/nequalstim/Desktop/'


    init_artist_profiles(ws)

    for row, artist in enumerate(artists, start=2):

        ws.cell(row=row, column=1).value = artist

        #viberate doesn't have lastfm listed, so we scrape it for all artists
        ws.cell(row=row, column=7).value = lastfm(artist)

        try:
            viberate(artist, ws, row)
        except:
            ws.cell(row=row, column=3).value = record_labels(artist)
            ws.cell(row=row, column=4).value = facebook(artist)
            ws.cell(row=row, column=6).value = instagram(artist)
            ws.cell(row=row, column=9).value = partyflock(artist)
            ws.cell(row=row, column=10).value= songkick(artist)

    wb.save(file_path + '/' + 'DJ_socials.xlsx')

