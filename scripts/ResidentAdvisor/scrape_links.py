import os
from bs4 import BeautifulSoup
from openpyxl import Workbook

from soup import get_soup, get_soup_js
from RA_links import active
from RA_site_elements import *
from headers import *


def scrape_links_venue(links, country):

    """
    The brain of the scraper. Takes in the links of venues and calls all individual scraper functions.
    Results will be written into a row of the excel file


    Arguments:
    links: array of residentadvisor links
    country: Name of country to be scraped

    Return:
    None

    Side Effects:
    Create Excel File ([country]_venues.xlsx) and fill it with all the scraped information.
    Creates a .txt file ([country]_events.txt), filled with RA events extracted from each venue. 
    """
    wb = Workbook()
    filename = country + '_venues.xlsx'
    ws = wb.active
    ws.title = country

    #initialize header in Excel sheet
    set_up_club(ws)

    file_path = os.getcwd

    row_count = 1
    events=[]

    for RA_link in links:
        site = get_soup_js(RA_link)

        #club had an event in 2016, 17, 18
        if active(site):
            row_count += 1

            ws.cell(row = row_count, column = 1).value = RA_link

            clubname = site.find('h1').string.encode('utf-8')
            ws.cell(row = row_count, column = 2).value = clubname

            #get picture links
            ws.cell(row = row_count, column = 3).value = picture(site)

            #get bio of venue
            ws.cell(row = row_count, column = 8).value = bio(site)

            #get address
            ws.cell(row = row_count, column = 14).value = address(site)

            try:
                top_bar = site.find('ul', {'class': 'clearfix'})
                #Get capacity number
                ws.cell(row = row_count, column = 13).value = capacity(top_bar)

                #Get contact number
                ws.cell(row=row_count, column = 7).value = phone(top_bar)
                
                sites = top_bar.find_all('a', href= True)

                #Get Website and Facebook
                website, facebook_page = homepage(sites)
                ws.cell(row = row_count, column = 9).value = website
                ws.cell(row = row_count, column = 20).value = facebook_page

                #get email
                ws.cell(row = row_count, column = 5).value = mail(sites)
            except: 
                pass

            #extract all the event links from venues
            get_events(site, events)

            # put a space in empty cells for nicer formatting in excel
            placeholders = [4, 6, 10, 11, 12, 15, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26]
            for column in placeholders:
                ws.cell(row = y, column = column).value = ' '
        else: 
            pass

    events_file = open(os.path.join(file_path, country+'_event.txt'), 'w')
    write_to_file(events, events_file)
    wb.save(file_path + '/' + filename)

def scrape_links_promoter(links, country):

    """
    The brain of the scraper. Takes in link and calls all individual scraper functions.

    Arguments:
    links: array of residentadvisor links
    country: Name of country to be scraped

    Return:
    None

    Side Effects:
    Create Excel File and fill it with all the scraped information
    Creates a .txt file ([country]_events.txt), filled with RA events extracted from each promoter. 

    """
    wb = Workbook()
    filename = country + '_promoters.xlsx'
    ws = wb.active
    ws.title = country

    #Initialize Excel header
    set_up_promoter(ws)

    file_path = os.getcwd()

    row_count = 1
    events=[]
    for link in links:
        site = get_soup_js(link)

        if active(site):
            #RA_link
            ws.cell(row = y, column = 3).value = link

            promoter_name = site.find('h1').string.encode('utf-8')
            ws.cell(row = y, column = 1).value = promoter_name

            #get picture links
            ws.cell(row = y, column = 7).value = picture(site)

            # print bio to file
            ws.cell(row = y, column = 8).value = bio(site)

            top_bar = site.find('ul', {'class': 'clearfix'})
            sites = top_bar.find_all('a', href= True)

            #Get Website and Facebook
            website, facebook_page = homepage(sites)
            ws.cell(row = y, column = 5).value = website
            ws.cell(row = y, column = 6).value = facebook_page

            #get email
            ws.cell(row = y, column = 10).value = mail(sites)

            placeholders = [2, 4, 9]
            for column in placeholders:
                ws.cell(row = y, column = column).value = ' '
            #extract all the event links from venues
            # put a space in empty cells for nicer formatting in excel
            get_events(site, events)
            events_filtered = list(set(events))

        else:
            pass

    events_file = open(os.path.join(file_path, country+'_event.txt'), 'w')
    write_to_file(events_filtered, events_file)
    #save excel file
    wb.save(file_path + '/' + filename)

def scrape_links_event(links):
    """
    This function scrapes all the information form RA event links. Links that have been collected
    from venues and promoters sites, which contain information such as eventname, lineup, promoter, venue etc. 

    Arguments:
    links: array filled with RA event link

    Return:
    None

    Side Effects:
    Write the collected information into Excel sheet (RA_events.xlsx)
    Output problem_artsist.txt file filled with artists, that had multiple matches
    in the Bookya DB

    """
    wb = Workbook()
    filename = 'RA_events.xlsx'
    ws = wb.active

    #Initialize Excel header
    set_up_event(ws)

    file_path = os.getcwd()

    #artists that have more than one match in the Bookya DB
    problem_artists = []

    for row, link in enumerate(links, start=2):

        site = get_soup(link)

        ws.cell(row=row, column=1).value = name(site)

        ws.cell(row=row, column = 4).value = 'noreply@bookya.com'

        div_event_item = site.find('div', {'id': 'event-item'})

        p_tags = div_event_item.find_all('p')

        #get line up
        notBookya, bookya, problem_artists = line_up(p_tags, problem_artists)
        ws.cell(row=row, column = 22).value = ','.join(notBookya)
        ws.cell(row=row, column = 23).value = ','.join(bookya)

        #get bio
        ws.cell(row=row, column=10).value = bio_event(p_tags)

        #get flyer picture
        ws.cell(row=row, column=3).value = flyer_picture(div_event_item)

        #dates and venue both need a_tags from top_bar
        top_bar = site.find('ul', {'class': 'clearfix'})
        a_tags = top_bar.find_all('a', href=True)

        start_date, end_date = dates(a_tags)
        ws.cell(row=row, column=13).value = start_date
        ws.cell(row=row, column=14).value = end_date

        found_venue, venue_name = venue(a_tags)
        if (found_venue):
            ws.cell(row=row, column=21).value = venue_name
        else:
            ws.cell(row=row, column=20).value = venue_name

        #get costs for event
        li = top_bar.find_all('li')
        cost, currency = costs(li)
        ws.cell(row=row, column = 15).value = cost
        ws.cell(row=row, column = 16).value = currency

        found_prom, promoter_name = promoter(li)
        if(found_prom):
            ws.cell(row=row, column=19).value = promoter_name
        else:
            ws.cell(row=row, column=18).value = promoter_name

        # put a space in empty cells for nicer formatting in excel
        placeholders = [2, 5, 6, 7, 8, 9, 11, 12, 17, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35]
        for column in placeholders:
            ws.cell(row = row, column = column).value = " "

    artistfile = open(os.path.join(file_path, 'problem_artists.txt'), 'w')
    problem_artists = list(set(problem_artists))
    write_to_file(problem_artists, artistfile)
    wb.save(file_path + '/' + filename)
