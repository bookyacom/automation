import os
from soup import get_soup_js
from socials import *
from openpyxl import Workbook
from links import artist_on_bookya
from bs4 import BeautifulSoup

partyflock_url = 'https://partyflock.nl'


def init_artist_profiles(ws):
    """
    Initialize Workbook

    Arguments: 
    ws: open workbook

    Side effects:
    Fill in first row of sheet

    Return:
    None
    """
    ws.cell(row=1, column = 1).value = 'display_name'
    ws.cell(row=1, column = 2).value = 'websites'
    ws.cell(row=1, column = 3).value = 'record_labels'
    ws.cell(row=1, column = 4).value = 'facebook_page'
    ws.cell(row=1, column = 5).value = 'beatport_dj'
    ws.cell(row=1, column = 6).value = 'instagram'
    ws.cell(row=1, column = 7).value = 'lastfm'
    ws.cell(row=1, column = 8).value = 'mixcloud'
    ws.cell(row=1, column = 9).value = 'partyflock'
    ws.cell(row=1, column = 10).value = 'songkick'
    ws.cell(row=1, column = 11).value = 'soundcloud'
    ws.cell(row=1, column = 12).value = 'spotify'
    ws.cell(row=1, column = 13).value = 'twitter'
    ws.cell(row=1, column = 14).value = 'youtube_channel'
    ws.cell(row=1, column = 15).value = 'bandsintown'
    ws.cell(row=1, column = 16).value = 'bio'
    ws.cell(row=1, column = 17).value = 'genres'
    ws.cell(row=1, column = 18).value = ' '

def artist_2017(artist_page):
    """
    Gives True or False to the question:
    Did this artist have an event in 2017 or 2018?

    Return:
    True: did have an event
    False: did not have an event
    """
    try:
        artist_agenda = artist_page.find('div', {'class': 'artist agenda box'})
        try:
            time_ = artist_agenda.find('time', datetime=True)
            time = str(time_['datetime'])
            if re.search('2018', time):
                return True
            else:
                if re.search('2017', time):
                    return True
                else:
                    return False
        except:
            date_ = artist_agenda.find_all('div', {'class': 'block'})
            date = date_[4].getText()
            if re.search('2018', date):
                return True
            else:
                if re.search('2017', date):
                    return True
                else:
                    return False
    except:
        return False

def get_information(artist_urls):
    """
    Organizer of this scraper. Takes in urls, gets the html +
    coordinates the Excel sheet, calls functions to fill in the blanks

    Arguments: 
    artist_urls: array filled with partyflock artist links
    e.g. ['/artist/39874:Dk', '/artist/38057:Delirium-Johnny']

    Side effects:
    Creates and initializes Workbook, gathers all the data (social, labels etc.)

    Return:
    None
    """

    wb = Workbook()
    filename = 'Partyflock_artists.xlsx'
    ws = wb.active
    ws.title = 'Partyflock'
    file_path = os.getcwd()

    init_artist_profiles(ws)
    row_count = 2
    for artist_url in artist_urls:
        artist_page = get_soup_js(partyflock_url + artist_url)
        if artist_2017(artist_page):
            artist = name(artist_page)
            if not artist_on_bookya(artist):
                ws.cell(row=row_count, column = 9).value = artist_url

                get_socials(artist_page, ws, row_count)

                ws.cell(row=row_count, column = 1).value = artist

                ws.cell(row=row_count, column = 2).value = website(artist_page)

                ws.cell(row=row_count, column = 3).value = labels(artist)

                ws.cell(row=row_count, column = 16).value = bio(artist_page)

                ws.cell(row=row_count, column = 17).value = genres(artist_page)

                # for nicer formatting in excel
                placeholders = [5, 7, 8, 10, 15, 18]
                for column in placeholders:
                    ws.cell(row=row_count, column=column).value = ' '

                row_count += 1

        else:
            pass
    wb.save(file_path + '/Partyflock.xlsx')
